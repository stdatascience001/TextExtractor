import os
import csv
import uuid
import html
from datetime import datetime, date
from typing import List, Dict, Any, Optional
import openpyxl

from services.document_parser.base_parser import BaseDocumentParser
from services.document_parser.models import ParsedDocumentWrapper, DocumentModel, PageInfo, BlockItem
from core.logging import logger

def get_column_letter(col_idx: int) -> str:
    """Convert a 1-based column index to Excel column letter (e.g. 1 -> A, 27 -> AA)."""
    letter = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        letter = chr(65 + remainder) + letter
    return letter

def normalize_cell_value(val) -> str:
    """Standardize dates, floats, and None values to consistent string representation."""
    if val is None:
        return ""
    if isinstance(val, (datetime, date)):
        return val.isoformat()
    if isinstance(val, float):
        if val.is_integer():
            return str(int(val))
        return str(val)
    return str(val).strip()

class ExcelParser(BaseDocumentParser):
    def parse(self, file_path: str, document_id: str) -> ParsedDocumentWrapper:
        logger.info(f"[ExcelParser] Initiating parsing for: {file_path}")
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            logger.error(f"[ExcelParser] Failed to load workbook: {str(e)}")
            raise e

        pages_list = []
        page_number = 1

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Detect merged cells and build a lookup map to flatten values
            merged_lookup = {}
            try:
                for merged_range in ws.merged_cells.ranges:
                    tl_val = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
                    for r in range(merged_range.min_row, merged_range.max_row + 1):
                        for c in range(merged_range.min_col, merged_range.max_col + 1):
                            merged_lookup[(r, c)] = tl_val
            except Exception as merge_err:
                logger.warning(f"[ExcelParser] Merged cell analysis skipped: {str(merge_err)}")

            # Read all rows in the sheet
            raw_rows = []
            max_r = ws.max_row or 0
            max_c = ws.max_column or 0
            
            # If the sheet is empty, check if we should skip or make placeholder
            if max_r == 0 or max_c == 0:
                logger.info(f"[ExcelParser] Worksheet {sheet_name} is empty. Skipping.")
                continue

            for r in range(1, max_r + 1):
                row_vals = []
                for c in range(1, max_c + 1):
                    val = merged_lookup.get((r, c)) if (r, c) in merged_lookup else ws.cell(row=r, column=c).value
                    row_vals.append(normalize_cell_value(val))
                raw_rows.append(row_vals)

            # Find the header row (first non-empty row)
            header_idx = -1
            for idx, row in enumerate(raw_rows):
                if any(cell != "" for cell in row):
                    header_idx = idx
                    break

            if header_idx == -1:
                # Fully empty sheet
                logger.info(f"[ExcelParser] Worksheet {sheet_name} has no non-empty rows. Skipping.")
                continue

            headers = raw_rows[header_idx]
            data_rows = raw_rows[header_idx + 1:]

            # Remove rows that are completely empty at the end
            while data_rows and not any(cell != "" for cell in data_rows[-1]):
                data_rows.pop()

            row_count = len(data_rows)
            col_count = len(headers)

            # Generate beautiful HTML table
            html_parts = ["<table class='spreadsheet-table'><thead><tr>"]
            for h in headers:
                html_parts.append(f"<th>{html.escape(h)}</th>")
            html_parts.append("</tr></thead><tbody>")
            
            for row in data_rows:
                html_parts.append("<tr>")
                for cell in row:
                    html_parts.append(f"<td>{html.escape(cell)}</td>")
                html_parts.append("</tr>")
            html_parts.append("</tbody></table>")
            table_html = "".join(html_parts)

            reading_order = 1
            items = []

            # 1. Sheet Header Element
            header_block_id = f"sheet-{page_number}-header"
            header_item = BlockItem(
                id=str(uuid.uuid4()),
                block_id=header_block_id,
                document_id=document_id,
                page_number=page_number,
                parent_block_id=None,
                type="sheet_header",
                text=f"Sheet: {sheet_name}",
                reading_order=reading_order,
                confidence=1.0,
                source_parser="openpyxl",
                metadata={
                    "sheet_name": sheet_name,
                    "row_count": row_count,
                    "col_count": col_count
                },
                children=[]
            )
            items.append(header_item)
            reading_order += 1

            # 2. Table Element
            table_block_id = f"sheet-{page_number}-table"
            
            # Create child table rows
            table_row_children = []
            for r_idx, row in enumerate(data_rows, start=header_idx + 2): # 1-indexed, relative to headers
                row_text = " | ".join(row)
                row_coord_cells = {
                    f"{get_column_letter(c_idx)}{r_idx}": val 
                    for c_idx, val in enumerate(row, start=1)
                }
                
                row_item = BlockItem(
                    id=str(uuid.uuid4()),
                    block_id=f"sheet-{page_number}-row-{r_idx}",
                    document_id=document_id,
                    page_number=page_number,
                    parent_block_id=table_block_id,
                    type="table_row",
                    text=row_text,
                    reading_order=reading_order,
                    confidence=1.0,
                    source_parser="openpyxl",
                    metadata={
                        "row_index": r_idx,
                        "cells": row_coord_cells
                    },
                    children=[]
                )
                table_row_children.append(row_item)
                reading_order += 1

            table_item = BlockItem(
                id=str(uuid.uuid4()),
                block_id=table_block_id,
                document_id=document_id,
                page_number=page_number,
                parent_block_id=None,
                type="table",
                text="",
                reading_order=reading_order,
                confidence=1.0,
                source_parser="openpyxl",
                table_html=table_html,
                metadata={
                    "headers": headers,
                    "rows": data_rows,
                    "table_html": table_html
                },
                children=table_row_children
            )
            items.append(table_item)
            reading_order += len(table_row_children) + 1

            pages_list.append(PageInfo(
                page_number=page_number,
                width=0.0,
                height=0.0,
                items=items
            ))
            page_number += 1

        if not pages_list:
            # If workbook has absolutely no valid data sheets, create an empty sheet placeholder
            placeholder_item = BlockItem(
                id=str(uuid.uuid4()),
                block_id="sheet-1-empty",
                document_id=document_id,
                page_number=1,
                parent_block_id=None,
                type="paragraph",
                text="Empty Worksheet Document",
                reading_order=1,
                confidence=1.0,
                source_parser="openpyxl",
                metadata={},
                children=[]
            )
            pages_list.append(PageInfo(
                page_number=1,
                width=0.0,
                height=0.0,
                items=[placeholder_item]
            ))

        return ParsedDocumentWrapper(
            document=DocumentModel(
                metadata={
                    "source_parser": "openpyxl",
                    "document_type": "spreadsheet",
                    "source_type": "upload",
                    "sheet_count": len(pages_list)
                },
                pages=pages_list
            )
        )

class CsvParser(BaseDocumentParser):
    def parse(self, file_path: str, document_id: str) -> ParsedDocumentWrapper:
        logger.info(f"[CsvParser] Initiating parsing for: {file_path}")
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        # Try to read file content with multiple encodings
        content = None
        encodings = ["utf-8", "utf-8-sig", "latin-1", "cp1252"]
        for enc in encodings:
            try:
                with open(file_path, "r", encoding=enc) as f:
                    content = f.read()
                break
            except UnicodeDecodeError:
                continue

        if content is None:
            raise UnicodeDecodeError("CSV parsing failed: Could not decode with common encodings.")

        # Detect delimiter using Sniffer
        delimiter = ","
        if content.strip():
            try:
                sniffer = csv.Sniffer()
                dialect = sniffer.sniff(content[:4000])
                delimiter = dialect.delimiter
            except Exception:
                delimiter = ","

        # Read CSV rows
        reader = csv.reader(content.splitlines(), delimiter=delimiter)
        raw_rows = [ [normalize_cell_value(val) for val in r] for r in reader if r ]

        # Parse rows
        if not raw_rows:
            placeholder_item = BlockItem(
                id=str(uuid.uuid4()),
                block_id="csv-empty",
                document_id=document_id,
                page_number=1,
                parent_block_id=None,
                type="paragraph",
                text="Empty CSV Document",
                reading_order=1,
                confidence=1.0,
                source_parser="csv",
                metadata={},
                children=[]
            )
            return ParsedDocumentWrapper(
                document=DocumentModel(
                    metadata={
                        "source_parser": "csv",
                        "document_type": "spreadsheet",
                        "source_type": "upload",
                        "sheet_count": 1
                    },
                    pages=[PageInfo(page_number=1, width=0.0, height=0.0, items=[placeholder_item])]
                )
            )

        headers = raw_rows[0]
        data_rows = raw_rows[1:]

        # Clean trailing empty rows
        while data_rows and not any(cell != "" for cell in data_rows[-1]):
            data_rows.pop()

        filename = os.path.basename(file_path)
        sheet_name = filename.split('.')[0]

        # HTML Table generation
        html_parts = ["<table class='spreadsheet-table'><thead><tr>"]
        for h in headers:
            html_parts.append(f"<th>{html.escape(h)}</th>")
        html_parts.append("</tr></thead><tbody>")
        
        for row in data_rows:
            html_parts.append("<tr>")
            for cell in row:
                html_parts.append(f"<td>{html.escape(cell)}</td>")
            html_parts.append("</tr>")
        html_parts.append("</tbody></table>")
        table_html = "".join(html_parts)

        reading_order = 1
        items = []

        # 1. Sheet Header Element
        header_item = BlockItem(
            id=str(uuid.uuid4()),
            block_id="sheet-1-header",
            document_id=document_id,
            page_number=1,
            parent_block_id=None,
            type="sheet_header",
            text=f"Sheet: {sheet_name}",
            reading_order=reading_order,
            confidence=1.0,
            source_parser="csv",
            metadata={
                "sheet_name": sheet_name,
                "row_count": len(data_rows),
                "col_count": len(headers)
            },
            children=[]
        )
        items.append(header_item)
        reading_order += 1

        # 2. Table Element
        table_block_id = "sheet-1-table"
        table_row_children = []
        for r_idx, row in enumerate(data_rows, start=2): # 1-indexed, header is row 1
            row_text = " | ".join(row)
            row_coord_cells = {
                f"{get_column_letter(c_idx)}{r_idx}": val 
                for c_idx, val in enumerate(row, start=1)
            }
            
            row_item = BlockItem(
                id=str(uuid.uuid4()),
                block_id=f"sheet-1-row-{r_idx}",
                document_id=document_id,
                page_number=1,
                parent_block_id=table_block_id,
                type="table_row",
                text=row_text,
                reading_order=reading_order,
                confidence=1.0,
                source_parser="csv",
                metadata={
                    "row_index": r_idx,
                    "cells": row_coord_cells
                },
                children=[]
            )
            table_row_children.append(row_item)
            reading_order += 1

        table_item = BlockItem(
            id=str(uuid.uuid4()),
            block_id=table_block_id,
            document_id=document_id,
            page_number=1,
            parent_block_id=None,
            type="table",
            text="",
            reading_order=reading_order,
            confidence=1.0,
            source_parser="csv",
            table_html=table_html,
            metadata={
                "headers": headers,
                "rows": data_rows,
                "table_html": table_html
            },
            children=table_row_children
        )
        items.append(table_item)

        return ParsedDocumentWrapper(
            document=DocumentModel(
                metadata={
                    "source_parser": "csv",
                    "document_type": "spreadsheet",
                    "source_type": "upload",
                    "sheet_count": 1
                },
                pages=[PageInfo(
                    page_number=1,
                    width=0.0,
                    height=0.0,
                    items=items
                )]
            )
        )
